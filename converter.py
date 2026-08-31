"""
Core parsing / cleaning / mapping logic for the CRM import converter.
No Flask here - this module is UI-agnostic so it can be tested on its own.

High level pipeline
-------------------
    read_table()          file  -> headers, rows, hyperlinks
    analyze_columns()     rows  -> per-column content statistics ("what is
                                   actually inside this column")
    detect_phone_columns()/scan_phones()
                          rows  -> phone columns exploded into sub-columns
    build_source_columns()      -> the list shown on the left of the mapping UI
    suggest_mapping()           -> pre-selected targets for that list
    build_records()       mapping -> output records + "misplaced value" report
    dedupe_records()      records -> merged records + duplicate report
    write_csv()

Two ideas drive most of the code:

1. *Atoms.* Every source cell is decomposed into typed atoms (emails, urls,
   phones, plain text) instead of being trusted to hold what its column
   header promises. An e-mail dropped into a phone column is still found,
   and is routed to an e-mail field instead of being thrown away.

2. *Duplicates are a view, not a decision.* The converter always produces
   BOTH the raw row-per-source-row table and the merged/deduplicated one,
   plus a report describing exactly which rows were merged, so a human can
   audit the merge instead of trusting it blindly.
"""
import csv
import io
import os
import re
import unicodedata
import zipfile
from urllib.parse import urlparse, parse_qs, unquote, urlunparse

import openpyxl

# ---------------------------------------------------------------------------
# Target CRM fields shown on the "right side" of the mapping screen.
# ---------------------------------------------------------------------------

TARGET_FIELDS = [
    {'id': 'company_lead', 'label': 'Company Name / Lead Name', 'type': 'text',
     'outputs': ['Company Name', 'Lead Name'],
     'hint': 'The value is written to 2 result columns at once: Company Name and Lead Name'},
    {'id': 'mobile_phone', 'label': 'Mobile', 'type': 'phone', 'outputs': ['Mobile'], 'hint': ''},
    {'id': 'home_phone', 'label': 'Home Phone', 'type': 'phone', 'outputs': ['Home Phone'], 'hint': ''},
    {'id': 'other_phone', 'label': 'Other Phone Number', 'type': 'phone', 'outputs': ['Other Phone Number'], 'hint': ''},
    {'id': 'corporate_website', 'label': 'Corporate Website', 'type': 'url', 'outputs': ['Corporate Website'], 'hint': ''},
    {'id': 'other_website', 'label': 'Other Website', 'type': 'url', 'outputs': ['Other Website'], 'hint': ''},
    {'id': 'work_email', 'label': 'Work E-mail', 'type': 'email', 'outputs': ['Work E-mail'], 'hint': ''},
    {'id': 'home_email', 'label': 'Home E-mail', 'type': 'email', 'outputs': ['Home E-mail'], 'hint': ''},
    {'id': 'other_email', 'label': 'Other E-mail', 'type': 'email', 'outputs': ['Other E-mail'], 'hint': ''},
    {'id': 'country_outreach', 'label': 'Country OUTREACH', 'type': 'text', 'outputs': ['Country OUTREACH'], 'hint': ''},
    {'id': 'outreach_comment', 'label': 'Outreach comment', 'type': 'text', 'outputs': ['Outreach comment'], 'hint': ''},
    {'id': 'comment', 'label': 'Comment', 'type': 'text', 'outputs': ['Comment'], 'hint': ''},
]
TARGET_BY_ID = {t['id']: t for t in TARGET_FIELDS}

OUTPUT_HEADERS = [
    'Company Name', 'Lead Name',
    'Mobile', 'Home Phone', 'Other Phone Number',
    'Corporate Website', 'Other Website',
    'Work E-mail', 'Home E-mail', 'Other E-mail',
    'Country OUTREACH', 'Outreach comment', 'Comment',
    'Source',
]

# Where a value of a given type is allowed to land, in priority order.
# Used both by the auto-mapping suggestions and by the "misplaced value"
# relocation logic.
TYPE_CHAINS = {
    'email': ['work_email', 'home_email', 'other_email'],
    'phone': ['mobile_phone', 'home_phone', 'other_phone'],
    'url': ['corporate_website', 'other_website'],
}

PHONE_OUTPUTS = ['Mobile', 'Home Phone', 'Other Phone Number']
EMAIL_OUTPUTS = ['Work E-mail', 'Home E-mail', 'Other E-mail']
URL_OUTPUTS = ['Corporate Website', 'Other Website']

SOURCE_VALUE = 'import'

# Max phone numbers pulled out of a single cell. Anything above this is
# counted in stats['overflow'] and surfaced in the UI as a warning rather
# than being dropped silently.
MAX_PHONES_PER_CELL = 5

# CSV dialect that matches the known-good reference_import.csv reference file:
# ';' delimiter, CRLF line endings, quote only when actually needed.
# (The legacy script used ',' + quote-all, which is why imports looked
# "crooked" - most locales/Bitrix read ';' as the field separator.)
CSV_DELIMITER = ';'
CSV_LINETERMINATOR = '\r\n'

# Default is UTF-8 *with* BOM. The BOM is the only in-band signal that tells
# Excel and most CRM importers "this is UTF-8". Without it they fall back to
# the system ANSI codepage, read the two UTF-8 bytes of "é" as two separate
# characters, and a later ASCII conversion turns each into '?' - which is
# where "México" -> "M??xico" (two question marks per letter) comes from.
CSV_ENCODING = 'utf-8-sig'

CSV_ENCODINGS = [
    {'id': 'utf-8-sig', 'label': 'UTF-8 with BOM (recommended)',
     'hint': 'Excel and most CRM importers detect the encoding automatically.'},
    {'id': 'utf-8', 'label': 'UTF-8 without BOM',
     'hint': 'Use if your importer trips over the BOM in the first header column.'},
    {'id': 'cp1252', 'label': 'Windows-1252 (Latin)',
     'hint': 'Use if the target system insists on a single-byte encoding. '
             'Spanish é, í, ñ survive; Cyrillic does not.'},
    {'id': 'cp1251', 'label': 'Windows-1251 (Cyrillic)',
     'hint': 'Only if the data contains Russian text and the importer expects ANSI.'},
]
CSV_ENCODING_IDS = {e['id'] for e in CSV_ENCODINGS}


# ---------------------------------------------------------------------------
# generic text cleaning
# ---------------------------------------------------------------------------

def clean_text(val):
    if val is None:
        return ''
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    val = unicodedata.normalize('NFKC', str(val)).strip()
    val = val.replace('\n', ' ').strip()
    val = re.sub(r'\s{2,}', ' ', val)
    return val


# ---------------------------------------------------------------------------
# email
# ---------------------------------------------------------------------------

EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')
# "user@gmail,com" -> "user@gmail.com" (comma typed instead of dot)
EMAIL_TYPO_RE = re.compile(r'(@[\w\-]+),([a-zA-Z]{2,4})\b')


def fix_email_typos(s):
    return EMAIL_TYPO_RE.sub(r'\1.\2', s)


def extract_emails(val):
    if not val:
        return []
    s = unicodedata.normalize('NFKC', str(val)).strip()
    s = fix_email_typos(s)
    s = s.replace('mailto:', ' ')
    found = EMAIL_RE.findall(s)
    seen = set()
    out = []
    for e in found:
        key = e.lower()
        if key not in seen:
            seen.add(key)
            out.append(e)
    return out


# ---------------------------------------------------------------------------
# phone
# ---------------------------------------------------------------------------

URL_RE = re.compile(r'https?://\S+')
WAME_RE = re.compile(r'wa\.me/\S+', re.IGNORECASE)
CTRL_ARTIFACT_RE = re.compile(r'(?:_x00[0-9A-Fa-f]{2}_)+')
# Straight and typographic single/double quotes.
QUOTE_RE = re.compile(r'[\'"‘’‚‛“”„‟]')

# Word-ish keywords. Matched with a boundary check (see looks_like_phone_header)
# so that "wa" does not fire on "Warehouse" and "tel" does not fire on "Hotel".
PHONE_HEADER_KEYWORDS = (
    'phone', 'phones', 'tel', 'telefono', 'teléfono', 'telefone', 'whatsapp',
    'wa', 'wsp', 'cel', 'celular', 'movil', 'móvil', 'mobile', 'number',
    'номер', 'телефон', 'моб', 'тел',
)
# Headers that must never be treated as phone columns even if their content
# happens to contain long digit runs. Long, unambiguous fragments are matched
# as substrings; short ones only as whole words, so that "Provider phone"
# (contains "id") is not vetoed.
NON_PHONE_HEADER_SUBSTRINGS = (
    'mail', 'correo', 'website', 'instagram', 'linkedin', 'facebook',
    'postal', 'fecha', 'дата', 'price', 'precio', 'цена', 'сайт', 'почт', 'ссыл',
)
NON_PHONE_HEADER_WORDS = {
    'web', 'site', 'url', 'link', 'id', 'zip', 'code', 'codigo', 'código', 'date',
}


def _split_parens(segment):
    extras = []

    def repl(m):
        content = m.group(1)
        if len(re.sub(r'\D', '', content)) >= 7:
            extras.append(content)
            return ' '
        return ' ' + content + ' '

    main = re.sub(r'\(([^)]*)\)', repl, segment)
    return main, extras


def _clean_phone_candidate(c):
    c = c.replace('(', '').replace(')', '')
    c = re.sub(r'^[^\d+]+', '', c)
    c = re.sub(r'[^\d\s+\-]+$', '', c)
    c = re.sub(r'\s{2,}', ' ', c).strip()
    if len(re.sub(r'\D', '', c)) < 5:
        return ''
    return c


def _split_phone_run(s):
    """Break a whitespace-joined run into separate phone numbers.

    People paste several numbers with nothing but a space between them
    ('+487765543 +36706709874'), while a *single* number is routinely
    written with spaces as digit-group separators ('+52 449 478 2400').
    A plain space can't tell those apart, so a new number is only started
    where something else marks it:
      - a '+' that isn't the first token - '+' opens a number, it never
        sits in the middle of one;
      - a '00' international-dialling prefix followed by enough digits to
        be a number on its own (not just a stray '00...' digit group);
      - a token that already looks like a complete number (>= 7 digits)
        showing up right after a group that is already complete-looking
        itself (>= MIN_INTERNATIONAL_DIGITS digits) - two full numbers
        typed back to back with no punctuation at all.
    Anything else (short digit groups such as '449', '478') is treated as
    part of the number currently being assembled.
    """
    tokens = [t for t in re.split(r'\s+', s.strip()) if t]
    out = []
    cur, cur_digits = [], 0
    for tok in tokens:
        tok_digits = len(re.sub(r'\D', '', tok))
        new_number = bool(cur) and (
            tok.startswith('+')
            or re.match(r'^00\d{7,}', tok)
            or (cur_digits >= MIN_INTERNATIONAL_DIGITS and tok_digits >= 7)
        )
        if new_number:
            out.append(' '.join(cur))
            cur, cur_digits = [tok], tok_digits
        else:
            cur.append(tok)
            cur_digits += tok_digits
    if cur:
        out.append(' '.join(cur))
    return out


def extract_phone_candidates(val):
    """Pull out every phone-like substring from a cell, cleaned but not yet
    '+'-normalized. Order preserved, de-duplicated by digit signature."""
    if not val:
        return []
    s = unicodedata.normalize('NFKC', str(val))
    s = CTRL_ARTIFACT_RE.sub('', s)
    # Quotes never belong to a phone number - only ever wrap or separate
    # one (copy-pasted from JSON/code). Turn them into a hard boundary
    # rather than leaving them stuck to a digit run.
    s = QUOTE_RE.sub(' ', s)

    phones = []
    for m in WAME_RE.findall(s):
        num = m.split('/', 1)[1] if '/' in m else ''
        num = re.sub(r'[^\d+]', '', num)
        if num:
            if not num.startswith('+'):
                num = '+' + num
            phones.append(num)

    s = WAME_RE.sub(' ', s)
    s = URL_RE.sub(' ', s)
    # an e-mail's local part can contain digits ("info2024@x.com") - drop
    # e-mails before looking for numbers
    s = EMAIL_RE.sub(' ', fix_email_typos(s))

    for part in re.split(r'[/&,;|]+', s):
        part = part.strip()
        if not part:
            continue
        main, extras = _split_parens(part)
        for cand in [main] + extras:
            for run in _split_phone_run(cand):
                cleaned = _clean_phone_candidate(run)
                if cleaned:
                    phones.append(cleaned)

    seen = set()
    out = []
    for p in phones:
        sig = re.sub(r'\D', '', p)
        if sig and sig not in seen:
            seen.add(sig)
            out.append(p)
    return out


def phone_signature(p):
    """Digits only - used for duplicate comparison."""
    return re.sub(r'\D', '', p or '')


# A national number anywhere in this data is at most 10 digits (Mexico,
# Colombia, Brazil, Spain, Russia all fit). From 11 digits up, the number
# almost certainly already carries its country code, so adding '+' is safe.
MAX_NATIONAL_DIGITS = 10
MIN_INTERNATIONAL_DIGITS = 11


def normalize_phone(raw, default_cc=None):
    """Ensure a phone string starts with '+'.

    Returns (value, status):
      'ok'       - already had a leading '+' (or 00 international prefix)
      'fixed'    - we added '+' to a number long enough to already carry a
                   country code (>= MIN_INTERNATIONAL_DIGITS)
      'cc_added' - we prepended the operator-supplied default country code
                   because the number looked like a national number
      'needs_cc' - a national number and no default country code was given.
                   Kept EXACTLY as it came, without a '+', and flagged.
      'invalid'  - too few digits to be a phone number at all; kept as-is.

    Why 'needs_cc' exists: prefixing '+' to a national number invents a country.
    '449 478 2400' is a local Mexican number, but '+449 478 2400' reads as
    country code 44 (United Kingdom) - a plausible-looking number that can never
    be dialled, which is worse than an obviously unformatted one. A number that
    cannot be resolved is left alone and reported instead of being guessed.
    """
    raw = (raw or '').strip()
    digits = re.sub(r'\D', '', raw)
    if not digits:
        return raw, 'invalid'
    if raw.startswith('+'):
        return raw, 'ok'
    if raw.startswith('00'):
        return '+' + raw[2:].strip(), 'ok'

    cc = re.sub(r'\D', '', default_cc or '')
    if cc and len(digits) <= MAX_NATIONAL_DIGITS and not digits.startswith(cc):
        # A leading 0 is a national trunk prefix ("098 554 7600" dialled
        # locally); it is never part of the international form.
        national = raw.lstrip().lstrip('0').lstrip() or raw
        return '+' + cc + ' ' + national, 'cc_added'
    if len(digits) >= MIN_INTERNATIONAL_DIGITS and not digits.startswith('0'):
        return '+' + raw, 'fixed'
    if len(digits) >= 7:
        return raw, 'needs_cc'
    return raw, 'invalid'


def _header_words(header):
    return set(re.split(r'[^0-9a-zа-яё]+', (header or '').lower()))


def looks_like_non_phone_header(header):
    h = (header or '').lower()
    if any(k in h for k in NON_PHONE_HEADER_SUBSTRINGS):
        return True
    return bool(_header_words(h) & NON_PHONE_HEADER_WORDS)


def looks_like_phone_header(header):
    h = (header or '').lower()
    if looks_like_non_phone_header(h):
        return False
    words = _header_words(h)
    if words & set(PHONE_HEADER_KEYWORDS):
        return True
    # allow "phone"/"telefono" as substrings of longer words, but never the
    # short ambiguous ones ('wa', 'tel', 'cel', ...)
    return any(k in h for k in ('phone', 'telefon', 'telefono', 'telefone', 'whatsapp', 'телефон'))


# ---------------------------------------------------------------------------
# urls (website)
# ---------------------------------------------------------------------------

TRACKING_PARAMS = {
    'utm_source', 'utm_medium', 'utm_content', 'utm_campaign',
    'fbclid', 'igshid', 'app_absent', 'text', 'brid', 'h', 'igsh',
    'utm_term', 'utm_id',
}

# Conservative TLD list: only used to decide whether a bare, scheme-less token
# such as "distribuidora-alpha.com.mx" is a website. Deliberately does NOT contain
# exotic ccTLDs, because Instagram handles in this dataset look like
# "alpha.cw" / "distribuidora.beta" and must stay plain text.
KNOWN_TLDS = {
    'com', 'net', 'org', 'info', 'biz', 'io', 'co', 'ai', 'app', 'dev', 'me',
    'shop', 'store', 'online', 'site', 'clinic', 'health', 'care', 'agency',
    'us', 'uk', 'eu', 'es', 'pt', 'fr', 'de', 'it', 'nl', 'pl', 'ru', 'ua',
    'mx', 'cl', 'ar', 'br', 'uy', 'py', 'bo', 'pe', 'ec', 've', 'pa', 'cr',
    'gt', 'hn', 'ni', 'sv', 'do', 'pr', 'cu', 'ca', 'au', 'nz', 'in', 'cn',
    'jp', 'kr', 'tr', 'il', 'ae', 'za',
}

URL_FULL_RE = re.compile(r'(?:https?://|www\.)[^\s,;<>"\']+', re.IGNORECASE)
BARE_DOMAIN_RE = re.compile(
    r'\b(?:[a-z0-9](?:[a-z0-9\-]{0,61}[a-z0-9])?\.)+([a-z]{2,10})\b(?:/[^\s,;<>"\']*)?',
    re.IGNORECASE,
)

# Hosts where the path identifies the account, so the path must be part of
# any duplicate key (otherwise every Instagram lead would merge into one).
SOCIAL_HOSTS = {
    'instagram.com', 'linkedin.com', 'facebook.com', 'fb.com', 'wa.me',
    't.me', 'twitter.com', 'x.com', 'tiktok.com', 'youtube.com',
}

# Links that are really a phone number in disguise. extract_phone_candidates
# turns them into numbers, so extract_urls must not also report them as
# websites (otherwise every "wa.me/5219..." lands in Corporate Website).
PHONE_LINK_HOSTS = {'wa.me', 'api.whatsapp.com', 'web.whatsapp.com'}


def extract_real_url(url):
    if not url:
        return ''
    url = url.strip()
    if url.startswith('mailto:'):
        return ''

    if 'l.facebook.com/l.php' in url:
        try:
            parsed = urlparse(url)
            params = parse_qs(parsed.query)
            if 'u' in params:
                return extract_real_url(params['u'][0])
        except Exception:
            pass

    if 'l.instagram.com' in url:
        try:
            parsed = urlparse(url)
            params = parse_qs(parsed.query)
            if 'u' in params:
                return extract_real_url(unquote(params['u'][0]))
        except Exception:
            pass

    try:
        parsed = urlparse(url)
        if parsed.query:
            params = parse_qs(parsed.query, keep_blank_values=True)
            cleaned = {k: v for k, v in params.items() if k not in TRACKING_PARAMS}
            new_query = '&'.join(f'{k}={v[0]}' for k, v in cleaned.items()) if cleaned else ''
            url = urlunparse(parsed._replace(query=new_query))
    except Exception:
        pass

    return url


def get_url(val, hyp_target=None):
    """Legacy single-value helper: treat the whole cell as one URL.
    Kept because it is the most permissive interpretation and is still used
    as a fallback for url-typed targets."""
    if hyp_target and hyp_target.startswith('mailto:'):
        hyp_target = None

    val_str = unicodedata.normalize('NFKC', str(val)).strip() if val is not None else ''

    url_to_use = None
    if val_str.startswith('http://') or val_str.startswith('https://'):
        url_to_use = val_str
    elif val_str and '.' in val_str and ' ' not in val_str and '@' not in val_str and len(val_str) > 4:
        url_to_use = hyp_target if hyp_target else ('https://' + val_str)
    elif hyp_target:
        url_to_use = hyp_target

    if not url_to_use:
        return ''

    return extract_real_url(url_to_use)


def _tidy_url(u):
    u = u.strip().rstrip('.,;:)')
    if u.lower().startswith('www.'):
        u = 'https://' + u
    elif not re.match(r'^https?://', u, re.IGNORECASE):
        u = 'https://' + u
    return extract_real_url(u)


def extract_urls(val, hyp_target=None):
    """Find every website-ish URL in a cell.

    Returns (confident, loose):
      confident - has a scheme / www. prefix / a recognised TLD. Safe to move
                  into a website field even if it came from another column.
      loose     - the whole-cell fallback (`get_url`) when nothing confident
                  was found. Used only when the target IS a website field.
    """
    s = unicodedata.normalize('NFKC', str(val)).strip() if val is not None else ''
    s = CTRL_ARTIFACT_RE.sub('', s)
    s = EMAIL_RE.sub(' ', fix_email_typos(s))

    found = []
    rest = s
    for m in URL_FULL_RE.findall(s):
        found.append(_tidy_url(m))
    rest = URL_FULL_RE.sub(' ', rest)

    for m in BARE_DOMAIN_RE.finditer(rest):
        if m.group(1).lower() in KNOWN_TLDS:
            found.append(_tidy_url(m.group(0)))

    if hyp_target and not hyp_target.startswith('mailto:'):
        found.append(extract_real_url(hyp_target))

    seen = set()
    confident = []
    for u in found:
        if not u:
            continue
        key = url_key(u)
        if not key or key.split('/')[0] in PHONE_LINK_HOSTS:
            continue  # it is a phone link - extract_phone_candidates owns it
        if key not in seen:
            seen.add(key)
            confident.append(u)

    loose = []
    if not confident:
        u = get_url(val, hyp_target)
        if u and url_key(u).split('/')[0] not in PHONE_LINK_HOSTS:
            loose.append(u)
    return confident, loose


def url_key(u):
    """Normalized key used for de-duplication and duplicate matching."""
    if not u:
        return ''
    try:
        p = urlparse(u if re.match(r'^https?://', u, re.I) else 'https://' + u)
    except Exception:
        return u.strip().lower()
    host = (p.netloc or '').lower().split(':')[0]
    if host.startswith('www.'):
        host = host[4:]
    path = (p.path or '').rstrip('/').lower()
    if host in SOCIAL_HOSTS:
        return host + path
    return host


def website_dup_key(u):
    """Key used when merging duplicates by website. Social links keep their
    path (the account), plain websites collapse to the bare domain."""
    key = url_key(u)
    if not key:
        return ''
    host = key.split('/')[0]
    if host in SOCIAL_HOSTS:
        # instagram.com/foo, linkedin.com/company/foo -> keep 2 segments max
        parts = key.split('/')
        return '/'.join(parts[:3]) if len(parts) > 2 else key
    return host


# ---------------------------------------------------------------------------
# value classification
# ---------------------------------------------------------------------------

def classify_value(val):
    """Best-effort type of a whole cell: 'email' | 'phone' | 'url' | 'text'
    | '' (empty). Used for column statistics and mapping suggestions."""
    s = clean_text(val)
    if not s:
        return ''
    if EMAIL_RE.search(fix_email_typos(s)):
        return 'email'
    confident, _loose = extract_urls(s)
    if confident:
        return 'url'
    digits = re.sub(r'\D', '', s)
    if len(digits) >= 7 and len(re.sub(r'[\d\s+()\-.,/xXeEtT]', '', s)) <= 2:
        return 'phone'
    if extract_phone_candidates(s):
        return 'phone'
    return 'text'


def extract_atoms(val, hyp_target=None, default_cc=None):
    """Decompose one raw cell into typed atoms.

    Returns dict:
      emails      list[str]
      urls        list[str]   confident websites
      urls_loose  list[str]   whole-cell fallback (see extract_urls)
      phones      list[str]   normalized, '+'-prefixed
      phone_status list[str]  status per phone, aligned with `phones`
      overflow    int         phones found beyond MAX_PHONES_PER_CELL
      text        str         cleaned whole cell
    """
    text = clean_text(val)
    emails = extract_emails(text)
    urls, urls_loose = extract_urls(text, hyp_target)
    cands = extract_phone_candidates(text)
    overflow = max(0, len(cands) - MAX_PHONES_PER_CELL)
    phones, statuses = [], []
    for c in cands[:MAX_PHONES_PER_CELL]:
        v, st = normalize_phone(c, default_cc)
        phones.append(v)
        statuses.append(st)
    return {
        'emails': emails,
        'urls': urls,
        'urls_loose': urls_loose,
        'phones': phones,
        'phone_status': statuses,
        'overflow': overflow,
        'text': text,
    }


# ---------------------------------------------------------------------------
# dedupe helpers used when several source columns map to the same target
# ---------------------------------------------------------------------------

def dedupe_join(values):
    seen = set()
    out = []
    for v in values:
        v = (v or '').strip()
        key = v.lower()
        if v and key not in seen:
            seen.add(key)
            out.append(v)
    return ', '.join(out)


def dedupe_join_urls(values):
    seen = set()
    out = []
    for v in values:
        v = (v or '').strip()
        if not v:
            continue
        key = url_key(v)
        if key not in seen:
            seen.add(key)
            out.append(v)
    return ', '.join(out)


def _merge_phone_list(values):
    """Return list of surviving phone strings, collapsing 'same number with /
    without country code' pairs."""
    kept = []  # [digits, original]
    for p in values:
        p = (p or '').strip()
        if not p:
            continue
        digits = phone_signature(p)
        if not digits:
            continue
        dup_idx = None
        for idx, (kd, _kp) in enumerate(kept):
            if kd == digits:
                dup_idx = idx
                break
            if kd.endswith(digits) and len(kd) - len(digits) <= 3:
                dup_idx = idx
                break
            if digits.endswith(kd) and len(digits) - len(kd) <= 3:
                kept[idx] = [digits, p]  # prefer the more complete version
                dup_idx = idx
                break
        if dup_idx is None:
            kept.append([digits, p])
    return [p for _, p in kept]


def dedupe_join_phones(values):
    return ', '.join(_merge_phone_list(values))


# ---------------------------------------------------------------------------
# reading the uploaded table
# ---------------------------------------------------------------------------

class TableTooLargeError(Exception):
    pass


MAX_ROWS = int(os.environ.get('MAX_ROWS', 20_000))
MAX_COLS = int(os.environ.get('MAX_COLS', 100))

# The real guard. A .xlsx is a ZIP: 1.4 MB of file can hold 42 MB of sheet XML,
# and openpyxl builds its whole object model before any row count is knowable,
# so MAX_ROWS alone cannot stop a memory blow-up - it only fires once the RAM is
# already spent. Measured on this parser: every MB of uncompressed XML costs
# ~0.42 s of CPU and ~5 MB of retained RAM, and that held from 42 MB (18 s /
# 212 MB) to 251 MB (106 s / 1293 MB). So one cap on the uncompressed size
# bounds CPU and memory at once - and closes the zip-bomb hole for free.
MAX_XLSX_UNCOMPRESSED = int(os.environ.get('MAX_XLSX_UNCOMPRESSED', 24 * 1024 * 1024))


def check_xlsx_archive(path):
    """Reject an .xlsx that would cost too much to parse, before parsing it.

    Reads only the ZIP central directory - no XML is touched, so a hostile
    file cannot attack the parser here."""
    try:
        with zipfile.ZipFile(path) as z:
            uncompressed = sum(i.file_size for i in z.infolist())
    except zipfile.BadZipFile:
        raise ValueError('Not a valid .xlsx file (it is not a ZIP archive).')
    if uncompressed > MAX_XLSX_UNCOMPRESSED:
        raise TableTooLargeError(
            f'The file unpacks to {uncompressed // (1024 * 1024)} MB of internal XML, '
            f'over the {MAX_XLSX_UNCOMPRESSED // (1024 * 1024)} MB limit. '
            f'Split it into several smaller files.')


def read_xlsx(path):
    check_xlsx_archive(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    # Only the active sheet is read. Silently dropping a populated sibling loses
    # data the operator believes they uploaded - but an empty leftover "Sheet2"
    # is normal in Excel and must not block anything, so only sheets that
    # actually hold rows count.
    other = [s.title for s in wb.worksheets
             if s is not ws and (s.max_row or 0) > 1]
    if other:
        raise TableTooLargeError(
            f'Only the active sheet ("{ws.title}") is imported, but these also '
            f'contain data: {", ".join(other)}. Leave one sheet in the file, or '
            f'upload each sheet separately.')
    max_col = ws.max_column or 0
    if max_col > MAX_COLS:
        raise TableTooLargeError(f'Too many columns: {max_col} (limit {MAX_COLS}).')
    if (ws.max_row or 0) > MAX_ROWS:
        raise TableTooLargeError(f'Too many rows: {ws.max_row} (limit {MAX_ROWS}).')
    headers = []
    for c in range(1, max_col + 1):
        h = clean_text(ws.cell(1, c).value)
        headers.append(h or f'Column {c}')

    rows = []
    hyperlinks = []
    for r in range(2, (ws.max_row or 1) + 1):
        raw_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if all(v is None or str(v).strip() == '' for v in raw_vals):
            continue
        rows.append([clean_text(v) for v in raw_vals])
        hyperlinks.append([
            (ws.cell(r, c).hyperlink.target if ws.cell(r, c).hyperlink else None)
            for c in range(1, max_col + 1)
        ])
    return headers, rows, hyperlinks


def _script_confusion(text):
    """How many words mix Latin and Cyrillic letters.

    That mixture is the signature of a single-byte codepage decoded with the
    wrong table: CP1252 bytes read as CP1251 turn "Clínica México" into
    "Clнnica Mйxico" - Cyrillic letters wedged inside Latin words. Real text is
    almost never mixed inside one word, so the count is ~0 for a correct
    decode and high for a wrong one."""
    bad = 0
    for word in re.findall(r'\w+', text):
        has_latin = any('a' <= c.lower() <= 'z' for c in word)
        has_cyrillic = any('Ѐ' <= c <= 'ӿ' for c in word)
        if has_latin and has_cyrillic:
            bad += 1
    return bad


def decode_csv_bytes(raw_bytes):
    """Decode CSV bytes, picking between the single-byte codepages by content.

    CP1251 and CP1252 cover the same byte range, so *both* always decode
    without raising - trying them in order silently mangles whichever one
    comes second. Spanish lead lists used to arrive as CP1252 and be read as
    CP1251. Ties keep the earlier candidate, so Cyrillic files behave as before.
    """
    for enc in ('utf-8-sig', 'utf-8'):
        try:
            return raw_bytes.decode(enc), enc
        except UnicodeDecodeError:
            continue
    best, best_enc, best_score = None, None, None
    for enc in ('cp1251', 'cp1252', 'latin-1'):
        try:
            candidate = raw_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
        score = _script_confusion(candidate)
        if best_score is None or score < best_score:
            best, best_enc, best_score = candidate, enc, score
        if best_score == 0:
            break
    if best is None:
        return raw_bytes.decode('utf-8', errors='replace'), 'utf-8'
    return best, best_enc


def read_csv_bytes(raw_bytes):
    text, _encoding = decode_csv_bytes(raw_bytes)

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=';,\t')
        delim = dialect.delimiter
    except Exception:
        delim = ';' if sample.count(';') >= sample.count(',') else ','

    reader = csv.reader(io.StringIO(text), delimiter=delim)
    data = [row for row in reader if any(cell.strip() for cell in row)]
    if not data:
        return [], [], []
    if len(data) > MAX_ROWS:
        raise TableTooLargeError(f'Too many rows: {len(data)} (limit {MAX_ROWS}).')
    if len(data[0]) > MAX_COLS:
        raise TableTooLargeError(f'Too many columns: {len(data[0])} (limit {MAX_COLS}).')
    headers = [clean_text(h) or f'Column {i + 1}' for i, h in enumerate(data[0])]
    width = len(headers)
    rows = []
    for r in data[1:]:
        r = list(r) + [''] * (width - len(r))
        rows.append([clean_text(v) for v in r[:width]])
    hyperlinks = [[None] * width for _ in rows]
    return headers, rows, hyperlinks


def _read_csv_file(path):
    with open(path, 'rb') as f:
        return read_csv_bytes(f.read())


def read_table(path, filename):
    lower = filename.lower()
    if lower.endswith('.csv') or lower.endswith('.txt'):
        return _read_csv_file(path)
    try:
        return read_xlsx(path)
    except TableTooLargeError:
        raise
    except ValueError:
        # Not a ZIP, so not really a spreadsheet. Operators do rename a CSV to
        # .xlsx; reading it as CSV is friendlier than "File is not a zip file".
        return _read_csv_file(path)


# ---------------------------------------------------------------------------
# column content analysis
# ---------------------------------------------------------------------------

HEADER_TARGET_HINTS = [
    (('company', 'empresa', 'compania', 'compañia', 'compañía', 'razon', 'razón',
      'name', 'nombre', 'lead', 'клиент', 'компан', 'название', 'имя'), 'company_lead'),
    (('mail', 'correo', 'почт'), 'email'),
    (('instagram', 'linkedin', 'facebook', 'social', 'red', 'соц'), 'social'),
    (('website', 'web', 'site', 'url', 'сайт', 'ссыл'), 'url'),
    (('phone', 'tel', 'whatsapp', 'wsp', 'cel', 'movil', 'móvil', 'mobile',
      'телефон', 'номер'), 'phone'),
    (('country', 'pais', 'país', 'страна'), 'country_outreach'),
    (('outreach',), 'outreach_comment'),
    (('comment', 'coment', 'note', 'nota', 'коммент', 'замет'), 'comment'),
]


def header_hint(header):
    h = (header or '').lower()
    for keywords, hint in HEADER_TARGET_HINTS:
        if any(k in h for k in keywords):
            return hint
    return ''


def analyze_columns(headers, rows, sample_size=400):
    """Per-column content statistics: how many non-empty cells look like an
    e-mail / a phone / a website / plain text, plus the dominant kind."""
    sample = rows[:sample_size]
    out = []
    for c, header in enumerate(headers):
        counts = {'email': 0, 'phone': 0, 'url': 0, 'text': 0}
        nonempty = 0
        for row in sample:
            cell = row[c] if c < len(row) else ''
            if not cell:
                continue
            nonempty += 1
            k = classify_value(cell)
            if k:
                counts[k] += 1
        dominant, share = '', 0.0
        if nonempty:
            dominant = max(counts, key=lambda k: counts[k])
            share = counts[dominant] / nonempty
            if counts[dominant] == 0:
                dominant, share = '', 0.0
        out.append({
            'index': c,
            'header': header,
            'nonempty': nonempty,
            'sampled': len(sample),
            'counts': counts,
            'dominant': dominant,
            'share': round(share, 3),
            'header_hint': header_hint(header),
        })
    return out


# ---------------------------------------------------------------------------
# phone-column detection + splitting ("scanning" step)
# ---------------------------------------------------------------------------

def detect_phone_columns(headers, rows, sample_size=400, col_stats=None):
    """A column is split into phone sub-columns when its header says "phone"
    OR its content is dominated by phone-shaped values.

    Content detection uses classify_value (whole-cell verdict) rather than
    "does any phone-like substring exist", so an e-mail/URL column is no
    longer mistaken for a phone column.
    """
    stats = col_stats if col_stats is not None else analyze_columns(headers, rows, sample_size)
    phone_cols = []
    for c, st in enumerate(stats):
        if looks_like_non_phone_header(headers[c]):
            continue
        header_hit = looks_like_phone_header(headers[c])
        nonempty = st['nonempty']
        content_hit = nonempty >= 1 and (st['counts']['phone'] / nonempty) >= 0.5
        if header_hit or content_hit:
            phone_cols.append(c)
    return phone_cols


def scan_phones(headers, rows, phone_col_indices, default_cc=None):
    """For every detected phone column, extract up to MAX_PHONES_PER_CELL
    normalized numbers per row, plus the non-phone leftovers (e-mails and
    websites accidentally typed into the phone column).

    Returns:
      phone_data: {col_idx: {'per_row': [[num, ...], ...],
                             'leftovers': [{'emails': [...], 'urls': [...]}, ...],
                             'max_count': int,
                             'stats': {'ok','fixed','cc_added','invalid','overflow',
                                       'foreign_emails','foreign_urls','empty'}}}
    """
    phone_data = {}
    for c in phone_col_indices:
        per_row = []
        leftovers = []
        stats = {'ok': 0, 'fixed': 0, 'cc_added': 0, 'needs_cc': 0, 'invalid': 0,
                 'overflow': 0, 'foreign_emails': 0, 'foreign_urls': 0, 'empty': 0}
        max_count = 1
        for row in rows:
            cell = row[c] if c < len(row) else ''
            atoms = extract_atoms(cell, None, default_cc)
            for st in atoms['phone_status']:
                stats[st] += 1
            stats['overflow'] += atoms['overflow']
            stats['foreign_emails'] += len(atoms['emails'])
            stats['foreign_urls'] += len(atoms['urls'])
            if cell and not atoms['phones']:
                stats['empty'] += 1
            per_row.append(atoms['phones'])
            leftovers.append({'emails': atoms['emails'], 'urls': atoms['urls']})
            max_count = max(max_count, len(atoms['phones']))
        phone_data[c] = {
            'per_row': per_row,
            'leftovers': leftovers,
            'max_count': min(max_count, MAX_PHONES_PER_CELL),
            'stats': stats,
        }
    return phone_data


def build_source_columns(headers, rows, phone_data, col_stats=None):
    """Build the flat list of mappable "source columns" shown in the UI:
    plain columns as-is, phone columns exploded into sub-columns."""
    stats = col_stats if col_stats is not None else analyze_columns(headers, rows)
    columns = []
    for c, header in enumerate(headers):
        st = stats[c]
        if c in phone_data:
            info = phone_data[c]
            n = info['max_count']
            for sub in range(n):
                key = f'{c}:{sub}'
                label = header if n == 1 else f'{header} — Phone {sub + 1}'
                samples = []
                for pr in info['per_row']:
                    if sub < len(pr) and pr[sub]:
                        samples.append(pr[sub])
                    if len(samples) >= 3:
                        break
                columns.append({
                    'key': key,
                    'label': label,
                    'kind': 'phone',
                    'samples': samples,
                    'stats': info['stats'] if sub == 0 else None,
                    'analysis': st if sub == 0 else None,
                })
        else:
            samples = []
            for row in rows:
                v = row[c] if c < len(row) else ''
                if v:
                    samples.append(v)
                if len(samples) >= 3:
                    break
            columns.append({
                'key': str(c),
                'label': header,
                'kind': st['dominant'] or 'text',
                'samples': samples,
                'stats': None,
                'analysis': st,
            })
    return columns


def suggest_mapping(source_columns):
    """Pre-fill the mapping dropdowns. Header wording wins; content type is
    the tie-breaker. Each concrete target is used at most once, spilling over
    to the next slot in its TYPE_CHAIN (Work -> Home -> Other e-mail, ...)."""
    used = set()
    suggestion = {}

    def take(chain, spill=True):
        """First unused slot of the chain. When the chain is exhausted and
        `spill` is on, reuse the last slot - two source columns mapped to the
        same target are joined with ', ', which is far better than silently
        dropping the 4th phone/e-mail column."""
        for t in chain:
            if t not in used:
                used.add(t)
                return t
        return chain[-1] if (spill and chain) else ''

    for col in source_columns:
        analysis = col.get('analysis') or {}
        hint = analysis.get('header_hint', '')
        kind = col.get('kind') or analysis.get('dominant') or ''
        target = ''
        if hint == 'company_lead':
            target = take(['company_lead'], spill=False)
        elif hint == 'email':
            target = take(TYPE_CHAINS['email'])
        elif hint == 'social':
            target = take(['other_website', 'corporate_website'])
        elif hint == 'url':
            target = take(TYPE_CHAINS['url'])
        elif hint == 'phone':
            target = take(TYPE_CHAINS['phone'])
        elif hint in ('country_outreach', 'outreach_comment', 'comment'):
            target = take([hint], spill=False)
        elif kind in TYPE_CHAINS:
            target = take(TYPE_CHAINS[kind])
        if target:
            suggestion[col['key']] = target
    return suggestion


# ---------------------------------------------------------------------------
# building the final output
# ---------------------------------------------------------------------------

DEFAULT_OPTIONS = {
    'autofix_types': True,       # move misplaced e-mails/phones/urls
    'default_country_code': '',  # e.g. '52' - used by normalize_phone
    'source_value': SOURCE_VALUE,
    'skip_rows_without_name': False,
    'sanitize_formulas': False,  # prefix =,+,-,@ with ' (Excel formula injection)
    'csv_encoding': CSV_ENCODING,
    'transliterate': False,      # é -> e, ñ -> n: guarantees pure ASCII output
    'dedupe_by_name': True,
    'dedupe_by_email': True,
    'dedupe_by_phone': True,
    'dedupe_by_website': False,
}


# A mappable source column is "<col>" or "<col>:<phone sub-column>".
SOURCE_KEY_RE = re.compile(r'^\d{1,4}(:\d{1,2})?$')

_BOOL_OPTIONS = ('autofix_types', 'skip_rows_without_name', 'sanitize_formulas',
                 'transliterate', 'dedupe_by_name', 'dedupe_by_email',
                 'dedupe_by_phone', 'dedupe_by_website')


def normalize_options(opts):
    """Coerce client-supplied options into the types the pipeline assumes.

    Everything here arrives as JSON from the browser, so nothing about its type
    is guaranteed: a list instead of an object used to raise AttributeError, and
    a dict in `source_value` reached csv writing and died there - both as a 500.
    """
    if opts is not None and not isinstance(opts, dict):
        raise ValueError('options must be an object.')
    o = dict(DEFAULT_OPTIONS)
    for k, v in (opts or {}).items():
        if k in o:
            o[k] = v
    for k in _BOOL_OPTIONS:
        o[k] = bool(o[k])
    o['default_country_code'] = re.sub(r'\D', '', str(o['default_country_code'] or ''))[:4]
    o['source_value'] = clean_text(o['source_value'])[:100] or SOURCE_VALUE
    if o['csv_encoding'] not in CSV_ENCODING_IDS:
        o['csv_encoding'] = CSV_ENCODING
    return o


def _text_for_key(key, row_idx, rows, phone_data):
    """Just the cell text for a source key - no atom extraction.

    A text-typed target uses nothing but `atoms['text']`, yet _atoms_for_key
    runs every e-mail/URL/phone regex first and the results are discarded
    (build_records skips the autofix block for text targets). Since rows[] is
    already clean_text'd by the reader, this collapses to one list index.
    Measured: 2.1x off build_records, byte-identical output."""
    if ':' in key:
        col_str, sub_str = key.split(':')
        c, sub = int(col_str), int(sub_str)
        info = phone_data.get(c)
        if not info:
            return None
        per_row = info['per_row']
        vals = per_row[row_idx] if row_idx < len(per_row) else []
        value = vals[sub] if sub < len(vals) else ''
        if value:
            return value
        if sub == 0 and row_idx < len(rows) and c < len(rows[row_idx]):
            return rows[row_idx][c]
        return ''
    c = int(key)
    row = rows[row_idx] if row_idx < len(rows) else []
    return row[c] if c < len(row) else ''


def _atoms_for_key(key, row_idx, rows, hyperlinks, phone_data, default_cc):
    """Resolve one source key on one row into atoms.

    For an exploded phone sub-column the phone value is already normalized by
    scan_phones; the non-phone leftovers of that cell are attached to sub 0
    so that they can still be recovered / relocated.
    """
    if ':' in key:
        col_str, sub_str = key.split(':')
        c, sub = int(col_str), int(sub_str)
        info = phone_data.get(c)
        if not info:
            return None
        per_row = info['per_row']
        vals = per_row[row_idx] if row_idx < len(per_row) else []
        value = vals[sub] if sub < len(vals) else ''
        left = info['leftovers'][row_idx] if row_idx < len(info['leftovers']) else {}
        raw_cell = rows[row_idx][c] if row_idx < len(rows) and c < len(rows[row_idx]) else ''
        return {
            'emails': left.get('emails', []) if sub == 0 else [],
            'urls': left.get('urls', []) if sub == 0 else [],
            'urls_loose': [],
            'phones': [value] if value else [],
            'phone_status': [],
            'overflow': 0,
            'text': value or (raw_cell if sub == 0 else ''),
        }
    c = int(key)
    row = rows[row_idx] if row_idx < len(rows) else []
    val = row[c] if c < len(row) else ''
    hyp = None
    if hyperlinks and row_idx < len(hyperlinks):
        hrow = hyperlinks[row_idx]
        hyp = hrow[c] if c < len(hrow) else None
    return extract_atoms(val, hyp, default_cc)


def _dup_key(kind, value):
    if kind == 'phone':
        return phone_signature(value)
    if kind == 'email':
        return (value or '').strip().lower()
    if kind == 'url':
        return url_key(value)
    return (value or '').strip().lower()


def build_records(headers, rows, hyperlinks, phone_data, mapping, options=None,
                  column_labels=None):
    """mapping: {source_key: target_id}.

    Returns (records, fixes):
      records - list of dicts keyed by OUTPUT_HEADERS, plus '_row' (1-based
                source row number, header row excluded) used by the reports.
      fixes   - list of dicts describing every value that was found in the
                "wrong" column, and what happened to it.
    """
    opts = normalize_options(options)
    labels = column_labels or {}
    autofix = bool(opts['autofix_types'])
    default_cc = opts['default_country_code']

    if mapping is not None and not isinstance(mapping, dict):
        raise ValueError('mapping must be an object of source key -> target id.')

    by_target = {}
    for key, target_id in (mapping or {}).items():
        if not target_id or target_id not in TARGET_BY_ID:
            continue
        if not SOURCE_KEY_RE.match(str(key)):
            # Keys come from the client. Without this, int() further down turns
            # 'abc', '1:2:3' or '-1' into a 500 - or, for '-1', silently reads
            # the last column instead of the one asked for.
            raise ValueError(f'Invalid source column key: {key!r}')
        by_target.setdefault(target_id, []).append(key)
    # stable, UI order
    for t in by_target:
        by_target[t].sort(key=lambda k: (int(k.split(':')[0]), int(k.split(':')[1]) if ':' in k else -1))

    records = []
    fixes = []

    for row_idx in range(len(rows)):
        buckets = {t: [] for t in by_target}
        pending = []  # (kind, value, from_target, source_label)

        for target_id, keys in by_target.items():
            ttype = TARGET_BY_ID[target_id]['type']
            for key in keys:
                if ttype == 'text':
                    text = _text_for_key(key, row_idx, rows, phone_data)
                    if text:
                        buckets[target_id].append(text)
                    continue

                atoms = _atoms_for_key(key, row_idx, rows, hyperlinks, phone_data, default_cc)
                if atoms is None:
                    continue
                label = labels.get(key, key)

                own = {'email': atoms['emails'],
                       'phone': atoms['phones'],
                       'url': atoms['urls'] + atoms['urls_loose']}[ttype]
                buckets[target_id].extend(own)

                if autofix:
                    for other in ('email', 'phone', 'url'):
                        if other == ttype:
                            continue
                        for v in {'email': atoms['emails'],
                                  'phone': atoms['phones'],
                                  'url': atoms['urls']}[other]:
                            pending.append((other, v, target_id, label))

        # ---- place relocated values -------------------------------------
        for kind, value, from_target, label in pending:
            chain = TYPE_CHAINS[kind]
            existing = set()
            for t in chain:
                for v in buckets.get(t, []):
                    existing.add(_dup_key(kind, v))
            if _dup_key(kind, value) in existing:
                fixes.append({'row': row_idx + 2, 'column': label, 'kind': kind,
                              'value': value, 'from': from_target, 'to': '',
                              'action': 'duplicate'})
                continue
            mapped = [t for t in chain if t in buckets]
            # Prefer a still-empty slot of the right type. When every slot is
            # taken, append to the LAST one ("Other Website" / "Other Phone
            # Number") - that is the natural overflow bucket, and it keeps the
            # primary field clean.
            candidates = ([t for t in mapped if not buckets[t]]
                          + [t for t in chain if t not in buckets]
                          + mapped[::-1]
                          + chain[::-1])
            dest = candidates[0]
            buckets.setdefault(dest, []).append(value)
            fixes.append({'row': row_idx + 2, 'column': label, 'kind': kind,
                          'value': value, 'from': from_target, 'to': dest,
                          'action': 'moved'})

        # ---- flatten into the output record ------------------------------
        record = {h: '' for h in OUTPUT_HEADERS}
        record['Source'] = opts['source_value']
        for target_id, values in buckets.items():
            if not values:
                continue
            ttype = TARGET_BY_ID[target_id]['type']
            if ttype == 'phone':
                joined = dedupe_join_phones(values)
            elif ttype == 'url':
                joined = dedupe_join_urls(values)
            else:
                joined = dedupe_join(values)
            for out_col in TARGET_BY_ID[target_id]['outputs']:
                record[out_col] = joined

        _strip_cross_field_duplicates(record)
        # +2 so the number matches what the operator sees in Excel
        # (row 1 is the header row).
        record['_row'] = row_idx + 2

        if opts['skip_rows_without_name'] and not record['Company Name']:
            continue
        records.append(record)

    return records, fixes


def _strip_cross_field_duplicates(record):
    """The same phone in Mobile and Home (or the same e-mail in Work and
    Other) is noise. Keep the first occurrence, drop the later ones."""
    for outs, kind, joiner in (
        (PHONE_OUTPUTS, 'phone', dedupe_join_phones),
        (EMAIL_OUTPUTS, 'email', dedupe_join),
        (URL_OUTPUTS, 'url', dedupe_join_urls),
    ):
        seen = set()
        for col in outs:
            if not record.get(col):
                continue
            kept = []
            for v in [x.strip() for x in record[col].split(',') if x.strip()]:
                k = _dup_key(kind, v)
                if k and k in seen:
                    continue
                seen.add(k)
                kept.append(v)
            record[col] = joiner(kept)


def _format_phone_cell(v):
    """Bitrix24's CRM import only binds a phone value when the cell carries
    a leading comma - the empty 'type' slot of its multi-value field format.
    A bare '+79223363661' is silently ignored on import; ',+79223363661' is
    read correctly. Everything upstream of this point (dedupe, cross-field
    duplicate stripping, merging) still works on the plain '+'-prefixed
    numbers joined by ', ' - this only reformats the string right before it
    is written to CSV. When more than one number ends up in the same output
    cell (once the dedicated Mobile/Home/Other Phone columns run out), each
    number keeps its own comma and they are separated by a plain space:
    ',+A ,+B ,+C'.
    """
    if not v:
        return v
    parts = [p.strip() for p in v.split(',') if p.strip()]
    return ' '.join(',' + p for p in parts)


def records_to_rows(records, sanitize_formulas=False, transliterate=False):
    rows = []
    for rec in records:
        row = []
        for h in OUTPUT_HEADERS:
            v = rec.get(h, '')
            if h in PHONE_OUTPUTS:
                v = _format_phone_cell(v)
            row.append(v)
        if transliterate:
            row = [strip_accents(v) for v in row]
        if sanitize_formulas:
            row = [_escape_formula(v) for v in row]
        rows.append(row)
    return rows


# "+52 33 1234 5678" starts with '+' but cannot be a formula - only digits and
# separators follow. Exempting it keeps phone numbers intact when formula
# protection is on, which is why that option used to be unusable for a CRM import.
PHONE_SHAPED_RE = re.compile(r'^\+[\d\s()\-.]+$')


def _escape_formula(v):
    if not isinstance(v, str):
        return v
    if v[:1] in ('=', '+', '-', '@', '\t', '\r') and not PHONE_SHAPED_RE.match(v):
        return "'" + v
    return v


# ---------------------------------------------------------------------------
# encoding safety net
# ---------------------------------------------------------------------------

# Characters with no useful NFKD decomposition that still deserve an ASCII
# spelling instead of '?'.
MANUAL_TRANSLIT = {
    'ß': 'ss', 'æ': 'ae', 'Æ': 'AE', 'ø': 'o', 'Ø': 'O', 'œ': 'oe', 'Œ': 'OE',
    'đ': 'd', 'Đ': 'D', 'ð': 'd', 'Ð': 'D', 'þ': 'th', 'Þ': 'TH',
    'ł': 'l', 'Ł': 'L', 'ª': 'a', 'º': 'o',
    '®': '(R)', '©': '(C)', '™': '(TM)', '€': 'EUR', '№': 'No',
    '–': '-', '—': '-', '−': '-', '‘': "'", '’': "'", '‚': "'",
    '“': '"', '”': '"', '„': '"', '…': '...', ' ': ' ',
}


def strip_accents(s):
    """é -> e, ñ -> n, ā -> a, ® -> (R).

    Latin letters only: Cyrillic is left alone (NFKD would turn 'й' into
    'и', which is a corruption, not a transliteration)."""
    if not s:
        return s
    out = []
    for ch in s:
        if ch in MANUAL_TRANSLIT:
            out.append(MANUAL_TRANSLIT[ch])
            continue
        decomposed = unicodedata.normalize('NFKD', ch)
        if len(decomposed) > 1 and decomposed[0].isascii():
            out.append(''.join(c for c in decomposed if not unicodedata.combining(c)))
        else:
            out.append(ch)
    return ''.join(out)


def fit_to_encoding(value, encoding):
    """Make `value` representable in `encoding` without raising.

    Falls back gracefully: exact character -> accent-stripped character ->
    '?'. Returns (value, replaced_count) so the UI can warn about losses."""
    if not value:
        return value, 0
    try:
        value.encode(encoding)
        return value, 0
    except UnicodeEncodeError:
        pass
    out = []
    replaced = 0
    for ch in value:
        try:
            ch.encode(encoding)
            out.append(ch)
            continue
        except UnicodeEncodeError:
            pass
        alt = strip_accents(ch)
        try:
            alt.encode(encoding)
            out.append(alt)
            replaced += 1
            continue
        except UnicodeEncodeError:
            pass
        out.append('?')
        replaced += 1
    return ''.join(out), replaced


# ---------------------------------------------------------------------------
# duplicate detection / merging
# ---------------------------------------------------------------------------

PAREN_RE = re.compile(r'^(.*?)\s*\(([^)]+)\)\s*$')


def alnum(s):
    return re.sub(r'[^a-z0-9]', '', (s or '').lower())


def name_keys(name):
    """"Distribuidora Alpha (distribuidora_alpha_mx)" produces both
    'distribuidoranautilus' and 'distribuidoranautilusmx', so the row merges
    with a plain 'distribuidora_alpha_mx' row."""
    name = (name or '').strip()
    if not name:
        return []
    m = PAREN_RE.match(name)
    if m:
        return [k for k in (alnum(m.group(1)), alnum(m.group(2))) if len(k) >= 3]
    k = alnum(name)
    return [k] if len(k) >= 3 else []


def pick_best_name(names):
    """Prefer a human-readable name ("Distribuidora Alpha (handle)") over a
    bare handle; longer wins as a tie-break."""
    names = [n for n in names if n]
    if not names:
        return ''
    return max(names, key=lambda n: (' ' in n, len(n)))


class _UnionFind:
    def __init__(self, n):
        self.parent = list(range(n))

    def find(self, i):
        while self.parent[i] != i:
            self.parent[i] = self.parent[self.parent[i]]
            i = self.parent[i]
        return i

    def union(self, i, j):
        ri, rj = self.find(i), self.find(j)
        if ri != rj:
            self.parent[max(ri, rj)] = min(ri, rj)


def _split_multi(value):
    return [v.strip() for v in (value or '').split(',') if v.strip()]


def record_dup_keys(rec, opts):
    """All duplicate keys of one record, tagged by kind so a phone can never
    collide with a company name."""
    keys = []
    if opts['dedupe_by_name']:
        for k in name_keys(rec.get('Company Name', '')):
            keys.append('n:' + k)
    if opts['dedupe_by_email']:
        for col in EMAIL_OUTPUTS:
            for v in _split_multi(rec.get(col, '')):
                keys.append('e:' + v.lower())
    if opts['dedupe_by_phone']:
        for col in PHONE_OUTPUTS:
            for v in _split_multi(rec.get(col, '')):
                d = phone_signature(v)
                if len(d) >= 8:
                    keys.append('p:' + d[-9:])
    if opts['dedupe_by_website']:
        for col in URL_OUTPUTS:
            for v in _split_multi(rec.get(col, '')):
                k = website_dup_key(v)
                if k:
                    keys.append('w:' + k)
    return keys


def dedupe_records(records, options=None):
    """Merge duplicate records.

    Returns (merged, groups):
      merged - list of records, in first-occurrence order
      groups - list of dicts describing every group that had >1 member:
               {'name', 'size', 'rows': [...], 'names': [...], 'matched_on': [...]}
    """
    opts = normalize_options(options)
    if not records:
        return [], []
    if not any(opts[k] for k in ('dedupe_by_name', 'dedupe_by_email',
                                 'dedupe_by_phone', 'dedupe_by_website')):
        return [dict(r) for r in records], []

    uf = _UnionFind(len(records))
    key_owner = {}
    for i, rec in enumerate(records):
        for k in record_dup_keys(rec, opts):
            if k in key_owner:
                uf.union(i, key_owner[k])
            else:
                key_owner[k] = i

    grouped = {}
    order = []
    for i in range(len(records)):
        root = uf.find(i)
        if root not in grouped:
            grouped[root] = []
            order.append(root)
        grouped[root].append(i)

    merged = []
    groups = []
    for root in order:
        members = [records[i] for i in grouped[root]]
        rec = _merge_group(members, opts)
        merged.append(rec)
        if len(members) > 1:
            groups.append({
                'name': rec['Company Name'],
                'size': len(members),
                'rows': [m.get('_row') for m in members],
                'names': [m.get('Company Name', '') for m in members],
                'matched_on': sorted(_infer_reasons(members, opts)),
            })
    return merged, groups


def _infer_reasons(members, opts):
    reasons = set()
    seen = {}
    for m in members:
        for k in record_dup_keys(m, opts):
            if k in seen and seen[k] is not m:
                reasons.add({'n': 'name', 'e': 'e-mail', 'p': 'phone', 'w': 'website'}[k[0]])
            seen[k] = m
    return reasons


# Only these hold ', '-joined lists of atoms. Everything else is free text,
# where a comma is punctuation: splitting "Call Monday, urgent" into two values
# and de-duplicating them against another row silently deletes words.
MULTI_VALUE_OUTPUTS = set(PHONE_OUTPUTS) | set(EMAIL_OUTPUTS) | set(URL_OUTPUTS)


def _merge_group(members, opts):
    rec = {h: '' for h in OUTPUT_HEADERS}
    best = pick_best_name([m.get('Company Name', '') for m in members])
    for h in OUTPUT_HEADERS:
        values = []
        for m in members:
            if h in MULTI_VALUE_OUTPUTS:
                values.extend(_split_multi(m.get(h, '')))
            else:
                v = (m.get(h) or '').strip()
                if v:
                    values.append(v)
        if h in ('Company Name', 'Lead Name'):
            rec[h] = best
        elif h in PHONE_OUTPUTS:
            rec[h] = dedupe_join_phones(values)
        elif h in URL_OUTPUTS:
            rec[h] = dedupe_join_urls(values)
        elif h == 'Source':
            rec[h] = next((v for v in values if v), opts['source_value'])
        else:
            rec[h] = dedupe_join(values)
    _strip_cross_field_duplicates(rec)
    rec['_row'] = members[0].get('_row')
    rec['_merged_from'] = [m.get('_row') for m in members]
    return rec


# ---------------------------------------------------------------------------
# reports
# ---------------------------------------------------------------------------

DUP_REPORT_HEADERS = ['Group', 'Kept Company Name', 'Merged rows',
                      'Source row', 'Source Company Name', 'Matched on']


def duplicates_report_rows(groups):
    rows = []
    for i, g in enumerate(groups, start=1):
        for row_no, name in zip(g['rows'], g['names']):
            rows.append([str(i), _escape_formula(g['name']), str(g['size']),
                         str(row_no), _escape_formula(name),
                         ', '.join(g['matched_on'])])
    return rows


FIX_REPORT_HEADERS = ['Source row', 'Source column', 'Detected type', 'Value',
                      'Mapped to', 'Moved to', 'Action']

_ACTION_LABEL = {'moved': 'moved', 'duplicate': 'duplicate, dropped'}
_KIND_LABEL = {'email': 'e-mail', 'phone': 'phone', 'url': 'website'}


def fixes_report_rows(fixes):
    rows = []
    for f in fixes:
        rows.append([
            str(f['row']), _escape_formula(f['column']),
            _KIND_LABEL.get(f['kind'], f['kind']),
            _escape_formula(f['value']),
            TARGET_BY_ID[f['from']]['label'] if f['from'] in TARGET_BY_ID else f['from'],
            TARGET_BY_ID[f['to']]['label'] if f['to'] in TARGET_BY_ID else '',
            _ACTION_LABEL.get(f['action'], f['action']),
        ])
    return rows


# ---------------------------------------------------------------------------
# writing
# ---------------------------------------------------------------------------

def write_csv(path, rows, headers=None, encoding=None):
    """Write the CSV and never raise on an unrepresentable character.

    Returns {'encoding', 'replaced', 'affected_rows'} so the caller can tell
    the operator how much was lost to the chosen codepage."""
    enc = encoding or CSV_ENCODING
    replaced = 0
    affected_rows = 0
    safe_rows = []
    for row in rows:
        row_replaced = 0
        safe = []
        for v in row:
            v, n = fit_to_encoding(v, enc)
            safe.append(v)
            row_replaced += n
        if row_replaced:
            affected_rows += 1
            replaced += row_replaced
        safe_rows.append(safe)

    # Write-then-rename: the download route serves this exact path, so writing
    # in place would let a client fetch a half-written file, or a mix of two
    # conversions. os.replace() is atomic within a filesystem.
    tmp_path = path + '.tmp'
    with open(tmp_path, 'w', newline='', encoding=enc) as f:
        writer = csv.writer(
            f, delimiter=CSV_DELIMITER, lineterminator=CSV_LINETERMINATOR,
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writerow(headers if headers is not None else OUTPUT_HEADERS)
        writer.writerows(safe_rows)
    os.chmod(tmp_path, 0o600)
    os.replace(tmp_path, path)
    return {'encoding': enc, 'replaced': replaced, 'affected_rows': affected_rows}


# ---------------------------------------------------------------------------
# backwards-compatible façade
# ---------------------------------------------------------------------------

def build_output_rows(headers, rows, hyperlinks, phone_data, mapping, options=None):
    """Old entry point: mapping -> list of OUTPUT_HEADERS rows, no dedupe."""
    records, _fixes = build_records(headers, rows, hyperlinks, phone_data, mapping, options)
    return records_to_rows(records)
